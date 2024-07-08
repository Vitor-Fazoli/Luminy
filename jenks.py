from openpyxl import load_workbook 
from openpyxl.styles import Color, PatternFill, Font, Border
from jenkspy import JenksNaturalBreaks

# Carrega o arquivo excel
workbook = load_workbook(filename='dados.xlsx')
sheet = workbook['principal']
indicate_column = 'B'
line = 2

# Gera a data de dados
data = []

# Recebe todas as informações da coluna indicada e armazena em uma data
while sheet[f"{indicate_column}{line}"].value is not None:
    data.append(sheet[f"{indicate_column}{line}"].value)
    line += 1

# Print de teste para verificar a data
print(data)

jnb_with_3_groups = JenksNaturalBreaks(3)
jnb_with_3_groups.fit(data) 

analysisSheet = workbook['analise']

#Padrões de cores
colorFills = [  
    PatternFill(start_color='D9FF0000', # Vermelho
                   end_color='D9FF0000',
                   fill_type='solid'),
    PatternFill(start_color='D98db4e2', # Azul
                   end_color='D98db4e2',
                   fill_type='solid'),
    PatternFill(start_color='D9fafc00', # Amarelo
                   end_color='D9fafc00',
                   fill_type='solid'),
    PatternFill(start_color='D900fe00', # Verde
                   end_color='D900fe00',
                   fill_type='solid'),
    PatternFill(start_color='D900fef3', # Ciano
                   end_color='D900fef3',
                   fill_type='solid'),
    PatternFill(start_color='D9ff00fe', # Magenta
                   end_color='D9ff00fe',
                   fill_type='solid')
]

colorIndex = 0
line = 3

for group in jnb_with_3_groups.groups_ :
    for value in group:
        analysisSheet[f"D{line}"].value = value
        analysisSheet[f"D{line}"].fill = colorFills[colorIndex]
        line += 1
    colorIndex += 1

workbook.save('dados.xlsx')